import random
from openpyxl import Workbook

# Define the work schedule matrix
schedule = [[0] * 5 for i in range(30)]

# Define a dictionary to keep track of the number of work days per employee
work_days = {"Alice": 0, "Bob": 0, "Charlie": 0, "David": 0, "Eve": 0}

# Define a list of employee names
names = list(work_days.keys())

# Populate the matrix with random work schedules
for i in range(30):
    # Check if all employees have already worked 16 days
    if all(v >= 16 for v in work_days.values()):
        print(f"Day {i+1}: All employees have already worked 16 days")
        break

    # Select a random employee who has worked less than 16 days so far
    available_employees = [name for name, days in work_days.items() if days < 16]
    print(f"Day {i+1}: Available employees = {available_employees}")
    if available_employees:
        employees = random.sample(available_employees, min(len(available_employees), 3))
        for j in employees:
            index = names.index(j)
            schedule[i][index] = 1
            work_days[j] += 1
        print(f"Day {i+1}: Assigned employees = {employees}")
    else:
        print(f"Day {i+1}: No available employees")

# Write the work schedule matrix to an Excel file
workbook = Workbook()
worksheet = workbook.active

# Write the employee names in the first row of the worksheet
worksheet.append([""] + names)

for i, row in enumerate(schedule):
    # Write the day number in the first column of each row
    worksheet.cell(row=i + 2, column=1, value=f"Day {i+1}")

    # Write the schedule for each employee in the row
    for j, value in enumerate(row):
        worksheet.cell(row=i + 2, column=j + 2, value=value)

workbook.save("work_schedule.xlsx")
